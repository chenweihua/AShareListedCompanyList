# -*- coding: gbk -*- 
'''
Created on 2016��4��20��

@author: shenyf
'''
import xlrd
from openpyxl.workbook import Workbook
import urllib
from syf.achievesselistedcompanyinfo import AchieveSSEStockInfo
from openpyxl.compat import range
from time import sleep

class ASharesListedCompaniesList:
    '''A�����й�˾�б�
    Ŀǰ�����ڽ��������������ġ����й�˾�б���ʽ�������⣬[����...]�����xls2xlsx ���������ڡ����й�˾�б���
    ׷���Ϻ����������й�˾��Ϣ���ǲ����ܵģ�ֻ�����ֶ��������ڡ����й�˾�б���Ȼ��ץȡ�Ϻ����й�˾��Ϣ��Ȼ���ֶ��ϲ�[������...]'''
    
    # file absolute path
    filePath = u"d:\\��֤��A�����й�˾�б�.xlsx"
    # �Ϻ����������й�˾��ʼ����Ϊ600000
    startStockNumber = 600000
    # xlsx ����
    __indexName = [u'��˾����', u'��˾���', u'��˾ȫ��', u'Ӣ������', u'ע���ַ', u'A�ɴ���', u'A�ɼ��', u'A����������', u'A���ܹɱ�', u'A����ͨ�ɱ�',
                   u'B�ɴ���', u'B�ɼ��', u'B����������', u'B���ܹɱ�', u'B����ͨ�ɱ�', u'�� ��', u'ʡ ��', u'�� ��', u'������ҵ', u'��˾��ַ']
    
    def appendSSEStocks(self, lastStockNumber):
        wb = Workbook()  # load_workbook(filename=self.filePath)
        sheet = wb.active
        
        # д����
        for i in range(self.__indexName.__len__()):
            _ = sheet.cell(column=i + 1, row=1, value=self.__indexName[i])
            
        row = sheet.max_row + 1
        for i in range(self.startStockNumber, lastStockNumber + 1):
            a = AchieveSSEStockInfo(i)
            sleep(1)
            
#             print a.getStatus()
            if not a.getStatus():
                continue
            
            for j in range(a.__public__.__len__()):
                m = a.__public__[j]
                f = getattr(a, m)
                print m
                print f()
                _ = sheet.cell(column=j + 1, row=row, value="%s" % f())
            row = row + 1
            
            # ÿ��ȡһ�����й�˾������Ϣ��д��xlsx������ռ�ù����ڴ�
            wb.save(filename=self.filePath)
    
    def downloadSZSEASharesListedCompaniesList(self):
        '''�������ڽ��������й�˾��Ϣ.'''
        dls = r"http://www.szse.cn/szseWeb/ShowReport.szse?SHOWTYPE=EXCEL&CATALOGID=1110&tab1PAGENUM=1&ENCODE=1&TABKEY=tab1"
        urllib.urlretrieve(dls, self.filePath)
        
    def storeASharesListedCompanies2XLS(self, lastStockNumber):
#         self.downloadSZSEASharesStocksInfo()
#         self.xls2xlsx()
        self.appendSSEStocks(lastStockNumber=lastStockNumber)

    def xls2xlsx(self):
        # first open using xlrd
        book = xlrd.open_workbook(self.filePath)
        index = 0
        nrows, ncols = 0, 0
        while nrows * ncols == 0:
            sheet = book.sheet_by_index(index)
            nrows = sheet.nrows
            ncols = sheet.ncols
            index += 1
    
        # prepare a xlsx sheet
        book1 = Workbook()
        sheet1 = book1.active
    
        for row in xrange(1, nrows):
            for col in xrange(1, ncols):
                sheet1.cell(row=row, column=col).value = sheet.cell_value(row, col)
        
        book1.save(filename=u'd:\\A.xlsx')
        return book1
        
if __name__ == '__main__':
    s = ASharesListedCompaniesList()
    s.storeASharesListedCompanies2XLS(lastStockNumber=600001)
